import pandas as pd
import inquirer
from openpyxl import load_workbook
imdb_df = pd.read_csv('data\IMDb movies.csv', low_memory=False)
from User import User
from preferences import Preferences

#ask user name and age
def get_know_user():
    #while True:
    user_name_input = input('What is your name?')
    name = user_name_input.strip() # remove spaces
    if name:
        name = name.capitalize() # name starts with capital
    age = int(input('What is your age?')) #age is integer

    user = User(name, age)
    return user

#age verification
def age_verification(user):
    if int(user.age)>=18:
        print("Welcome ", user.name, "! Thanks for providing your age (",
                  user.age,"). You are above the legal age, so your movie recommendation will be adjusted accordingly")
    else:
        print("Welcome ", user.name, "! Thanks for providing your age (",
                  user.age, "). You are below the legal age, so your movie recommendation will be adjusted accordingly")


def get_preferences(user):
    global ordering
    years = input("What year?")
    genres = input("What genre would you like to watch?")
    languages = input("What should be the original language?")
    ordering = ordering_func()
    preferences = Preferences(user.name, languages, years, genres, ordering)
    return preferences


def get_log(user, preferences):
    file_in_out = 'log.xlsx'
    wb = load_workbook(filename=file_in_out)
    ws = wb['log']
    last_row = ws.max_row
    for col_idx, val in enumerate([user.name,user.age,preferences.genres,preferences.languages,preferences.years,get_recommendation(imdb_df, preferences).sort_values(by=ordering, ascending=False).head(1)['original_title'].to_string(),rent]):
        ws.cell(column=col_idx + 1, row=last_row + 1, value=val)
    wb.save(file_in_out)


# gets called if user inputs one or more genres
def get_genre(data_df, genres):
    genre_list = str(genres).split()
    genre_rec = data_df[data_df['genre'].isin(genre_list)]
    return genre_rec

# gets called if user inputs one or more languages
def get_language(data_df, languages):
    language_list = str(languages).split()
    lang_rec = data_df[data_df['language'].isin(language_list)]
    return lang_rec


# gets called if user inputs one or more years
def get_year(data_df, years):
    year_list = str(years).split()
    year_rec = data_df[data_df['year'].isin(year_list)]
    return year_rec


# gets called when user is asked based on which variable they want to filter their movie
def  ordering_func():
    questions = [
        inquirer.List('sort',
                      message="What variable do you want to sort the answers on?",
                      choices=['Duration (longest)', 'Budget', 'Gross income (USA)'],
                      ),
    ]
    answers = inquirer.prompt(questions)
    ordering_var = answers["sort"]
    if answers["sort"]== "Duration (longest)":
        ordering_var = 'duration'
    elif answers["sort"]== "Budget":
        ordering_var = 'budget'
    else:
        ordering_var = 'usa_gross_income'

    return ordering_var

def get_recommendation(rec_df, preferences): # add argument
    recommendation = []
    if preferences.genres:
        rec_df = get_genre(rec_df, preferences.genres)
    if preferences.languages:
        rec_df = get_language(rec_df, preferences.languages)
    if preferences.years:
        rec_df = get_year(rec_df, preferences.years)

    # rec = rec_df.head(1)['original_title'] # if we want only the title
    rec_df = rec_df.head(10)[['original_title', 'year', 'genre', 'language', 'budget', 'duration','usa_gross_income', 'reviews_from_users']] # check that the filtering worked

    # check that the dataframe is not empty, if it is add a message suggesting user to try again with less strict options
    if rec_df.empty:
        print("The dataframe is empty, try again with less filters or maybe more options")
    else:
        print( "I found {} recommendation for you. Here's the one with the highest critics rating: ".format(len(rec_df.index)))
        #return rec_df.sort_values('year', ascending=False).head(1)[['original_title', 'year', 'genre', 'language']]
        print(rec_df.sort_values(by=ordering, ascending=False).head(1)[['original_title', 'year', 'genre', 'language']])
    return rec_df


def rent_option():
    global rent
    questions = [
        inquirer.List('rent',
                      message="Would you like to rent this movie?",
                      choices=['Yes', 'No'],
                      ),
    ]
    answers = inquirer.prompt(questions)
    rent = answers["rent"]
    if answers["rent"]== "Yes":
        print("Great! That costs 1,5 EUR")
    else:
        print("Nevermind. Come back later!")


def main():

    print("Movie recommendation program")
    user = get_know_user()
    age_verification(user)
    print("Let's figure out your movie preferences")
    preferences = get_preferences(user)
    recommendation = get_recommendation(imdb_df, preferences)
    print("Here is our recommendation: ", recommendation)
    rent_option()
    get_log(user, preferences)



if __name__ == "__main__":
    main()